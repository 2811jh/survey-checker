"""
generate_report.py — 根据问卷检查结果生成 Excel 报告

使用方式:
    python generate_report.py --input check_result.json --output 问卷检查报告.xlsx
"""

import argparse
import json
import sys
from datetime import datetime
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("❌ 请先安装 openpyxl: pip install openpyxl")
    sys.exit(1)


# 样式定义
HEADER_FONT = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
TITLE_FONT = Font(name="微软雅黑", size=14, bold=True)
NORMAL_FONT = Font(name="微软雅黑", size=10)
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

# 严重程度颜色
SEVERITY_COLORS = {
    "高": PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid"),
    "中": PatternFill(start_color="FFD93D", end_color="FFD93D", fill_type="solid"),
    "低": PatternFill(start_color="6BCB77", end_color="6BCB77", fill_type="solid"),
}

STATUS_COLORS = {
    "✅ 良好": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
    "⚠️ 建议改进": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
    "❌ 需要修改": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
}


def style_header_row(ws, row_num: int, col_count: int):
    """给表头行添加样式"""
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER


def style_data_cell(cell, wrap=True):
    """给数据单元格添加基本样式"""
    cell.font = NORMAL_FONT
    cell.alignment = Alignment(vertical="center", wrap_text=wrap)
    cell.border = THIN_BORDER


def auto_width(ws, min_width=10, max_width=50):
    """自动调整列宽"""
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                cell_len = len(str(cell.value or ""))
                # 中文字符按2个宽度计算
                cjk_count = sum(1 for c in str(cell.value or "") if '\u4e00' <= c <= '\u9fff')
                cell_len = cell_len + cjk_count
                if cell_len > max_length:
                    max_length = cell_len
            except Exception:
                pass
        adjusted = min(max(max_length + 2, min_width), max_width)
        ws.column_dimensions[col_letter].width = adjusted


def create_overview_sheet(wb, check_result: dict):
    """创建总览 Sheet"""
    ws = wb.active
    ws.title = "总览"
    
    # 标题
    ws.merge_cells("A1:B1")
    title_cell = ws["A1"]
    title_cell.value = "📋 问卷质量检查报告"
    title_cell.font = TITLE_FONT
    title_cell.alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 40
    
    # 概览信息
    overview_data = [
        ("问卷名称", check_result.get("survey_name", "未知")),
        ("检查时间", check_result.get("check_time", datetime.now().strftime("%Y-%m-%d %H:%M"))),
        ("题目总数", check_result.get("total_questions", 0)),
        ("发现问题总数", check_result.get("total_issues", 0)),
        ("错别字问题数", check_result.get("typo_count", 0)),
        ("逻辑问题数", check_result.get("logic_count", 0)),
        ("设计建议数", check_result.get("design_count", 0)),
        ("整体评分", f'{check_result.get("overall_score", 0)}/100'),
    ]
    
    headers = ["项目", "内容"]
    for col_idx, header in enumerate(headers, 1):
        ws.cell(row=3, column=col_idx, value=header)
    style_header_row(ws, 3, 2)
    
    for row_idx, (label, value) in enumerate(overview_data, 4):
        label_cell = ws.cell(row=row_idx, column=1, value=label)
        value_cell = ws.cell(row=row_idx, column=2, value=value)
        style_data_cell(label_cell)
        style_data_cell(value_cell)
        label_cell.font = Font(name="微软雅黑", size=10, bold=True)
    
    auto_width(ws)
    return ws


def create_typo_sheet(wb, typos: list):
    """创建错别字检查 Sheet"""
    ws = wb.create_sheet("错别字检查")
    
    headers = ["题号", "位置", "原始文本", "错误内容", "修改建议", "严重程度"]
    for col_idx, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_idx, value=header)
    style_header_row(ws, 1, len(headers))
    
    for row_idx, typo in enumerate(typos, 2):
        data = [
            typo.get("question_index", ""),
            typo.get("location", ""),
            typo.get("original_text", ""),
            typo.get("error_content", ""),
            typo.get("suggestion", ""),
            typo.get("severity", "中"),
        ]
        for col_idx, value in enumerate(data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            style_data_cell(cell)
            # 严重程度着色
            if col_idx == 6 and value in SEVERITY_COLORS:
                cell.fill = SEVERITY_COLORS[value]
    
    if not typos:
        ws.merge_cells("A2:F2")
        cell = ws.cell(row=2, column=1, value="✅ 未发现错别字问题")
        style_data_cell(cell)
        cell.alignment = Alignment(horizontal="center")
    
    auto_width(ws)
    return ws


def create_logic_sheet(wb, logic_issues: list):
    """创建逻辑检查 Sheet"""
    ws = wb.create_sheet("逻辑检查")
    
    headers = ["题号", "问题类型", "问题描述", "修改建议", "严重程度"]
    for col_idx, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_idx, value=header)
    style_header_row(ws, 1, len(headers))
    
    for row_idx, issue in enumerate(logic_issues, 2):
        data = [
            issue.get("question_index", ""),
            issue.get("issue_type", ""),
            issue.get("description", ""),
            issue.get("suggestion", ""),
            issue.get("severity", "中"),
        ]
        for col_idx, value in enumerate(data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            style_data_cell(cell)
            if col_idx == 5 and value in SEVERITY_COLORS:
                cell.fill = SEVERITY_COLORS[value]
    
    if not logic_issues:
        ws.merge_cells("A2:E2")
        cell = ws.cell(row=2, column=1, value="✅ 未发现逻辑问题")
        style_data_cell(cell)
        cell.alignment = Alignment(horizontal="center")
    
    auto_width(ws)
    return ws


def create_design_sheet(wb, design_evals: list):
    """创建设计专业性 Sheet"""
    ws = wb.create_sheet("设计专业性")
    
    headers = ["评估维度", "状态", "说明", "改进建议"]
    for col_idx, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_idx, value=header)
    style_header_row(ws, 1, len(headers))
    
    for row_idx, eval_item in enumerate(design_evals, 2):
        data = [
            eval_item.get("dimension", ""),
            eval_item.get("status", ""),
            eval_item.get("description", ""),
            eval_item.get("suggestion", ""),
        ]
        for col_idx, value in enumerate(data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            style_data_cell(cell)
            if col_idx == 2 and value in STATUS_COLORS:
                cell.fill = STATUS_COLORS[value]
    
    if not design_evals:
        ws.merge_cells("A2:D2")
        cell = ws.cell(row=2, column=1, value="✅ 整体设计良好")
        style_data_cell(cell)
        cell.alignment = Alignment(horizontal="center")
    
    auto_width(ws)
    return ws


def create_original_sheet(wb, questions: list):
    """创建问卷原文 Sheet"""
    ws = wb.create_sheet("问卷原文")
    
    headers = ["题号", "题型", "题目内容", "选项/说明", "是否必填", "跳转逻辑"]
    for col_idx, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_idx, value=header)
    style_header_row(ws, 1, len(headers))
    
    for row_idx, q in enumerate(questions, 2):
        options_text = "\n".join(q.get("options", [])) if q.get("options") else ""
        data = [
            q.get("index", row_idx - 1),
            q.get("type", "未知"),
            q.get("title", ""),
            options_text,
            "是" if q.get("required") else "否",
            q.get("logic", "") or "",
        ]
        for col_idx, value in enumerate(data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            style_data_cell(cell)
    
    auto_width(ws)
    return ws


def generate_report(check_result: dict, output_path: str):
    """生成完整的 Excel 报告"""
    wb = Workbook()
    
    # 创建各个 Sheet
    create_overview_sheet(wb, check_result)
    create_typo_sheet(wb, check_result.get("typos", []))
    create_logic_sheet(wb, check_result.get("logic_issues", []))
    create_design_sheet(wb, check_result.get("design_evals", []))
    create_original_sheet(wb, check_result.get("questions", []))
    
    # 保存
    wb.save(output_path)
    print(f"✅ 报告已保存到: {output_path}")
    return output_path


def main():
    parser = argparse.ArgumentParser(description="生成问卷检查 Excel 报告")
    parser.add_argument("--input", type=str, required=True, help="检查结果 JSON 文件路径")
    parser.add_argument("--output", type=str, default=None, help="输出 Excel 文件路径")
    
    args = parser.parse_args()
    
    with open(args.input, "r", encoding="utf-8") as f:
        check_result = json.load(f)
    
    if not args.output:
        survey_name = check_result.get("survey_name", "未知问卷")
        date_str = datetime.now().strftime("%Y%m%d")
        args.output = f"问卷检查报告_{survey_name}_{date_str}.xlsx"
    
    generate_report(check_result, args.output)


if __name__ == "__main__":
    main()
